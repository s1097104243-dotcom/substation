import io
import re
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook


def normalize_header(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip().replace(" ", "").lower()


def find_plan_index(row_values) -> int:
    for i, cell in enumerate(row_values):
        normalized = normalize_header(cell)
        if normalized == "plan" or normalized.startswith("plan"):
            return i
    return -1


def extract_structured_data(df: pd.DataFrame) -> list[dict]:
    structured_data = []
    current_plan_title = "未知Plan"
    current_test_title = "未知測試項目"
    header_row_index = -1
    headers = []
    header_indices = {}

    for index, row in df.iterrows():
        row_str = row.astype(str).tolist()

        plan_index = find_plan_index(row.tolist())
        if plan_index != -1:
            plan_text = row.iloc[8] if len(row) > 8 else None
            if pd.isna(plan_text):
                for j in range(plan_index + 1, len(row)):
                    if pd.notna(row.iloc[j]):
                        plan_text = row.iloc[j]
                        break
            current_plan_title = str(plan_text).strip() if pd.notna(plan_text) else "未知Plan"
            current_test_title = "未知測試項目"
            header_row_index = -1
            headers = []
            header_indices = {}

        if any(cell.strip().lower() == "test" for cell in row_str):
            try:
                test_index = next(i for i, cell in enumerate(row_str) if cell.strip().lower() == "test")
                title_text = row.iloc[9] if len(row) > 9 else None
                if pd.isna(title_text):
                    for j in range(test_index + 1, len(row)):
                        if pd.notna(row.iloc[j]):
                            title_text = row.iloc[j]
                            break
                current_test_title = str(title_text).strip() if pd.notna(title_text) else "未知測試項目"
                header_row_index = -1
                headers = []
                header_indices = {}
            except Exception:
                continue

        if any("label" in cell.lower() for cell in row_str):
            header_row_index = index
            headers = [str(c).strip() for c in row.tolist()]
            header_indices = {}
            for i, cell in enumerate(row.tolist()):
                normalized = normalize_header(cell)
                if normalized in ("label/value", "actual", "result", "%error"):
                    header_indices[normalized] = i
            continue

        if header_row_index != -1 and any(pd.notna(cell) for cell in row):
            if any(cell.strip().lower() == "test" for cell in row_str):
                header_row_index = -1
                headers = []
                header_indices = {}
                continue

            data_values = row.tolist()
            if len(headers) == len(data_values) and "label/value" in header_indices:
                data_dict = {"Plan": current_plan_title, "Test": current_test_title}
                data_dict["Label/Value"] = data_values[header_indices["label/value"]]
                if "actual" in header_indices:
                    data_dict["Actual"] = data_values[header_indices["actual"]]
                if "result" in header_indices:
                    data_dict["Result"] = data_values[header_indices["result"]]
                if "%error" in header_indices:
                    data_dict["% Error"] = data_values[header_indices["%error"]]
                structured_data.append(data_dict)
            elif len(headers) != len(data_values):
                header_row_index = -1

    return structured_data


def build_data_maps(cleaned_df: pd.DataFrame) -> dict:
    all_data_maps = {}
    plan_groups = cleaned_df.groupby("Plan", sort=False)

    for plan_name, plan_df in plan_groups:
        data_map = {}
        for _, row in plan_df.iterrows():
            try:
                test_name = str(row["Test"]).strip()
                if "復閉" in test_name or "始動" in test_name or "瞬跳" in test_name:
                    key = test_name
                else:
                    raw = str(row["Label/Value"]).strip().split()[0]
                    val = float(raw)
                    suffix = str(int(val)) if val == int(val) else f"{val:.2f}".rstrip("0")
                    key = f"{test_name}{suffix}"
                data_map[key] = row["Actual"]
            except Exception:
                continue

        sheet_name = str(plan_name).strip()
        all_data_maps[sheet_name] = data_map

    return all_data_maps


def process_workbook(source_files, target_file):
    all_structured_data = []
    parse_logs = []

    for src in source_files:
        try:
            src.seek(0)
            df = pd.read_excel(src, header=None)
            structured_data = extract_structured_data(df)
            if not structured_data:
                parse_logs.append(f"⚠️ {src.name}: 未擷取到資料，已略過")
                continue
            all_structured_data.extend(structured_data)
            parse_logs.append(f"✅ {src.name}: 擷取 {len(structured_data)} 筆")
        except Exception as exc:
            parse_logs.append(f"❌ {src.name}: 讀取失敗（{exc}）")

    final_df = pd.DataFrame(all_structured_data)
    if final_df.empty:
        raise ValueError("未能成功擷取任何來源資料，請檢查 Excel 格式")

    keep_cols = ["Plan", "Test", "Label/Value", "Actual", "Result", "% Error"]
    existing_cols = [c for c in keep_cols if c in final_df.columns]
    cleaned_df = final_df[existing_cols].copy()

    if "Actual" in cleaned_df.columns:
        cleaned_df.loc[:, "Actual"] = cleaned_df.loc[:, "Actual"].apply(
            lambda x: re.sub(r"[sva]", "", str(x), flags=re.IGNORECASE) if pd.notna(x) else x
        )

    all_data_maps = build_data_maps(cleaned_df)
    if not all_data_maps:
        raise ValueError("沒有可用的 Plan 對應資料")

    target_file.seek(0)
    suffix = Path(target_file.name).suffix.lower()
    keep_vba = suffix == ".xlsm"

    wb_write = load_workbook(target_file, keep_vba=keep_vba)
    target_file.seek(0)
    wb_values = load_workbook(target_file, data_only=True, keep_vba=keep_vba)

    matched_count = 0
    skipped_count = 0
    write_logs = []

    for ws in wb_write.worksheets:
        if not ws.title.startswith("特性_"):
            continue

        ws_values = wb_values[ws.title]
        o1_value = ws_values["O1"].value
        if o1_value is None:
            o1_value = ws["O1"].value

        if o1_value is None:
            skipped_count += 1
            write_logs.append(f"--- 跳過 [{ws.title}]：O1 為空")
            continue

        panel_name = str(o1_value).strip()
        if panel_name not in all_data_maps:
            skipped_count += 1
            write_logs.append(f"--- 跳過 [{ws.title}]：找不到盤名 [{panel_name}]")
            continue

        current_map = all_data_maps[panel_name]
        write_count = 0
        for r in range(1, ws.max_row + 1):
            label = ws[f"CN{r}"].value
            if label is not None and label in current_map:
                ws[f"CO{r}"].value = current_map[label]
                write_count += 1

        matched_count += 1
        write_logs.append(f">>> [{ws.title}] 盤名 [{panel_name}]：寫入 {write_count} 筆")

    output = io.BytesIO()
    wb_write.save(output)
    output.seek(0)

    return {
        "file_bytes": output.getvalue(),
        "cleaned_df": cleaned_df,
        "parse_logs": parse_logs,
        "write_logs": write_logs,
        "matched_count": matched_count,
        "skipped_count": skipped_count,
        "plan_count": len(all_data_maps),
    }

def render_template_downloads() -> None:
    st.subheader("檔案下載")

    assets = [
        (
            Path("assets") / "公版_IED試驗報告.xlsm",
            "下載試驗報告",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
        (
            Path("assets") / "公版_變電所_測試程序.psx",
            "下載測試程序",
            "application/octet-stream",
        ),
    ]

    for path, label, mime in assets:
        if path.exists():
            with path.open("rb") as f:
                st.download_button(
                    label=label,
                    data=f.read(),
                    file_name=path.name,
                    mime=mime,
                    use_container_width=True,
                )
        else:
            st.info(f"尚未找到 `{path}`")







st.set_page_config(page_title="Protection Substation Tool", layout="wide")
st.title("Protection 變電所報告轉換")
st.caption("上傳來源報告（可多個）與目標檔案，產生可下載的更新結果。")
render_template_downloads()
st.divider()

source_uploads = st.file_uploader(
    "來源資料檔案（可多選）",
    type=["xlsx", "xlsm", "xls"],
    accept_multiple_files=True,
)
target_upload = st.file_uploader(
    "目標檔案（建議 xlsm / xlsx）",
    type=["xlsx", "xlsm"],
    accept_multiple_files=False,
)

if st.button("開始轉換", type="primary", use_container_width=True):
    if not source_uploads:
        st.error("請至少上傳 1 個來源檔案。")
    elif target_upload is None:
        st.error("請上傳目標檔案。")
    else:
        with st.spinner("處理中，請稍候..."):
            try:
                result = process_workbook(source_uploads, target_upload)
            except Exception as exc:
                st.exception(exc)
            else:
                st.success("處理完成！")
                c1, c2, c3 = st.columns(3)
                c1.metric("成功匹配分頁", result["matched_count"])
                c2.metric("跳過分頁", result["skipped_count"])
                c3.metric("盤名數", result["plan_count"])

                st.subheader("來源解析紀錄")
                st.code("\n".join(result["parse_logs"]) or "無")

                st.subheader("寫入紀錄")
                st.code("\n".join(result["write_logs"]) or "無")

                st.subheader("擷取資料預覽")
                st.dataframe(result["cleaned_df"].head(200), use_container_width=True)

                out_name = f"updated_{Path(target_upload.name).name}"
                st.download_button(
                    "下載更新後檔案",
                    data=result["file_bytes"],
                    file_name=out_name,
                    mime="application/vnd.ms-excel",
                    use_container_width=True,
                )

